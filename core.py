from Auths import gmail, password
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time


class MyBlog:
    def __init__(self):
        self.options = webdriver.ChromeOptions()
        self.driver = webdriver.Chrome(options=self.options)

    def auth_pro_culture(self):
        driver = self.driver

        driver.get('https://pro.culture.ru/new/auth/login')
        time.sleep(3)
        driver.find_element(By.CLASS_NAME, "ant-input").send_keys(gmail)
        time.sleep(1)
        driver.find_elements(By.CLASS_NAME, 'ant-input')[1].send_keys(password)
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR,
                            ".ant-btn.ant-btn-primary").click()
        time.sleep(4)
        driver.find_element(By.CSS_SELECTOR, ".main-form_organization-link").click()
        time.sleep(3)

    def is_my_blog(self):
        driver = self.driver

        source = r"D:\Codes\Check_blog\organizations - 2022-06-20T042354.449.xlsx"
        wb = openpyxl.load_workbook(source)  # path to the Excel file
        ws = wb.active

        ws[f"C1"] = 'Может писать в блог ?'
        count_of_orgs = 28059
        begin_of_read = 2
        for i in range(begin_of_read, count_of_orgs + 2):
            org_id = ws[f'A{i}'].value
            print(f"A{i}", org_id)
            try:
                driver.get(f'https://pro.culture.ru/new/subordinate/organizations/edit/{org_id}')
            except Exception as ex:
                ws[f'C{org_id}'] = 'Не открылось'
                continue
            # time.sleep(2)
            try:
                # driver.find_element(By.CSS_SELECTOR, ".ant-btn.ant-btn-primary").click()
                time.sleep(3)
                # elements = driver.find_elements(By.CSS_SELECTOR, ".ant-col.ant-col-12")[9].text
                # driver.find_element(By.CSS_SELECTOR, ".ant-checkbox.ant-checkbox-checked")
                element = driver.find_element(By.XPATH,
                                              "/html/body/div[1]/div[1]/form/div/div[3]/div[1]/div/"
                                              "div/div/div[2]/div/div/div/div[10]/div/div/div/div[1]/label")

                element.find_element(By.CSS_SELECTOR, ".ant-checkbox.ant-checkbox-checked")
                ws[f"C{i}"] = "Да"
                print(f"Есть блог!")
            except Exception as ex:
                print(ex)
                print("Нет блога!")
                ws[f"C{i}"] = "Нет"
            if i % 100 == 0:
                wb.save(source)

        wb.save(source)
        driver.quit()
        driver.close()
